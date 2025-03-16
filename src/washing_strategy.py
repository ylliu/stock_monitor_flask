from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from src.tushare_interface import TushareInterface

file = open('stock_data.txt', 'w')

# 假设参数
H_max = 300  # 假设的最大横盘天数
R_max = 15  # 假设的最大涨停次数
V_max = 800  # 假设的最大流通市值（例如100亿）

# 权重分配（示例）
w_H = 0.4
w_R = 0.4
w_V = 0.2


def normalize(value, max_value):
    return value / max_value if max_value > 0 else 0


class RealInfo:
    def __init__(self, code, name, price, change, limit_circ_mv, free_circ_mv, is_low_ma5, is_low_ma10, start_date,
                 end_date, concept, max_turnover_rate, angle_of_5, angle_of_10, angle_of_20, angle_of_30):
        self.code = code
        self.name = name
        self.price = price
        self.change = change
        self.limit_circ_mv = limit_circ_mv
        self.free_circ_mv = free_circ_mv
        self.is_low_ma5 = is_low_ma5
        self.is_low_ma10 = is_low_ma10
        self.start_date = start_date
        self.end_date = end_date
        self.concept = concept
        self.max_turnover_rate = max_turnover_rate
        self.angle_of_5 = angle_of_5
        self.angle_of_10 = angle_of_10
        self.angle_of_20 = angle_of_20
        self.angle_of_30 = angle_of_30


class SearchResult:
    def __init__(self, code, name, count, start_date, end_date, limit_circ_mv, free_circ_mv,
                 concept, max_turnover_rate, angle_of_5, angle_of_10, angle_of_20, angle_of_30):
        self.code = code
        self.name = name
        self.count = count
        self.start_date = start_date
        self.end_date = end_date
        self.limit_circ_mv = limit_circ_mv
        self.free_circ_mv = free_circ_mv
        self.concept = concept
        self.max_turnover_rate = max_turnover_rate
        self.angle_of_5 = angle_of_5
        self.angle_of_10 = angle_of_10
        self.angle_of_20 = angle_of_20
        self.angle_of_30 = angle_of_30

    def __eq__(self, other):
        if not isinstance(other, SearchResult):
            return False
        return (self.code, self.name, self.start_date, self.end_date, ...) == (
            other.code, other.name, self.start_date, self.end_date, ...)  # 根据需要添加更多属性

    def __hash__(self):
        return hash((self.code, self.name, self.start_date, self.end_date, ...))  # 根据需要添加更多属性

    def __repr__(self):
        # 使用类名和所有属性的值来构建字符串
        return f"{self.__class__.__name__}(code={self.code!r}, name={self.name!r}, count={self.count}, start_date={self.start_date!r}, " \
               f"end_date={self.end_date!r}, limit_circ_mv={self.limit_circ_mv!r}, limit_circ_mv={self.limit_circ_mv!r}, concept={self.concept!r})"


class WashingStrategyConfig:
    def __init__(self, back_days, end_date, enable_local_run,
                 volume_rate, positive_average_pct, second_positive_high_days, before_positive_limit_circ_mv_min,
                 before_positive_limit_circ_mv_max, before_positive_free_circ_mv_min, before_positive_free_circ_mv_max,
                 positive_to_ten_mean_periods, ten_mean_scaling_factor, min_positive_days, is_margin_stock,
                 max_volume_high_days, five_days_max_up_pct, ten_days_max_up_pct, is_second_day_price_up):
        self.back_days = back_days
        self.end_date = end_date
        self.enable_local_run = enable_local_run
        self.volume_rate = volume_rate
        self.positive_average_pct = positive_average_pct
        self.second_positive_high_days = second_positive_high_days
        self.before_positive_limit_circ_mv_min = before_positive_limit_circ_mv_min
        self.before_positive_limit_circ_mv_max = before_positive_limit_circ_mv_max
        self.before_positive_free_circ_mv_min = before_positive_free_circ_mv_min
        self.before_positive_free_circ_mv_max = before_positive_free_circ_mv_max
        self.positive_to_ten_mean_periods = positive_to_ten_mean_periods
        self.ten_mean_scaling_factor = ten_mean_scaling_factor
        self.min_positive_days = min_positive_days
        self.is_margin_stock = is_margin_stock
        self.max_volume_high_days = max_volume_high_days
        self.five_days_max_up_pct = five_days_max_up_pct
        self.ten_days_max_up_pct = ten_days_max_up_pct
        self.is_second_day_price_up = is_second_day_price_up


class WashingStrategy:
    def __init__(self, stock_list, end_date, back_days, debug_mode, data_interface, config):
        self.daily_lines = None
        self.low_price_list = None
        self.search_results = []
        self.stock_list = stock_list
        self.end_date = end_date
        self.back_days = back_days
        self.count = 0
        self.debug_mode = debug_mode
        self.data_interface = data_interface
        self.config = config
        self.realtime_daily_lines = {}

    def set_daily_lines(self, daily_lines):
        self.daily_lines = daily_lines
        self.low_price_list = []
        for line in daily_lines:
            self.low_price_list.append(line.low)

    def search(self):
        previous_daily_line = self.daily_lines[0]
        # 分两步：第一步找到所有连阳的
        max_vol = 0
        max_vol_day = None
        max_pct_change = 0
        up_shadow_pct_of_max_pct_change_day = 0
        max_high_price = 0
        up_max_pct = 0
        min_low_price = 99999
        count = 0
        max_turnover_rate = 0
        for index_pos in range(1, len(self.daily_lines)):
            day = self.daily_lines[index_pos]
            first_positive_day = self.daily_lines[1]
            # 第二天不需要上涨的配置
            if index_pos == 2 and not self.config.is_second_day_price_up:
                if first_positive_day.is_volume_increased(previous_daily_line.vol, self.config.volume_rate):
                    count += 1
                    min_low_price = min(min_low_price, day.low)
                    max_high_price = max(max_high_price, day.high)
                    max_turnover_rate = max(max_turnover_rate, day.turnover_rate)
                    max_vol = max(max_vol, day.vol)
                    if day.max_pct_change > max_pct_change:
                        up_shadow_pct_of_max_pct_change_day = day.up_shadow_pct
                    continue

            if first_positive_day.is_volume_increased(previous_daily_line.vol, self.config.volume_rate) \
                    and day.is_higher_than_yesterday_close_price(self.daily_lines[index_pos - 1].close):
                count += 1
                min_low_price = min(min_low_price, day.low)
                max_high_price = max(max_high_price, day.high)
                max_turnover_rate = max(max_turnover_rate, day.turnover_rate)
                max_vol = max(max_vol, day.vol)
                if day.max_pct_change > max_pct_change:
                    up_shadow_pct_of_max_pct_change_day = day.up_shadow_pct
                continue

            else:
                break

        if count < self.config.min_positive_days:
            return None

        limit_circ_mv = self.data_interface.get_circ_mv3(day.code, previous_daily_line.trade_date[:10].replace("-", ""))
        if limit_circ_mv is not None:
            if limit_circ_mv > self.config.before_positive_limit_circ_mv_max or limit_circ_mv < self.config.before_positive_limit_circ_mv_min:
                # print('circ_mv bigger than 30:', circ_mv)
                return None
        free_circ_mv = self.data_interface.get_circ_mv4(day.code, previous_daily_line.trade_date[:10].replace("-", ""))
        if free_circ_mv is not None:
            if free_circ_mv > self.config.before_positive_free_circ_mv_max or free_circ_mv < self.config.before_positive_free_circ_mv_min:
                # print('circ_mv bigger than 30:', circ_mv)
                return None
        if limit_circ_mv is None or free_circ_mv is None:
            return None
        pct1 = (self.daily_lines[1].high - self.daily_lines[1].low) / previous_daily_line.close * 100
        pct2 = (self.daily_lines[2].high - self.daily_lines[2].low) / previous_daily_line.close * 100
        average_pct = round((pct1 + pct2) / 2, 2)
        if self.config.min_positive_days > 2 and self.daily_lines[3].is_positive():
            pct3 = (self.daily_lines[3].high - self.daily_lines[3].low) / previous_daily_line.close * 100
            average_pct = round((pct1 + pct2 + pct3) / 3, 2)

        # print(average_pct)
        if average_pct < self.config.positive_average_pct:
            return None

        start_date = self.daily_lines[1].trade_date[:10].replace("-", "")
        if count + 1 == len(self.daily_lines):
            end_date_pos = index_pos
        else:
            end_date_pos = index_pos - 1
        end_date = self.daily_lines[end_date_pos].trade_date[:10].replace("-", "")

        second_positive_high_days = self.config.second_positive_high_days
        is_max_high = self.data_interface.is_break_days_high(day.code,
                                                             self.daily_lines[end_date_pos].trade_date[:10].replace(
                                                                 "-", ""), second_positive_high_days)
        if not is_max_high:
            return None

        # 最大的成交量是xx天内的新高
        is_vol_max_high = self.data_interface.is_vol_break_days_high(day.code,
                                                                     end_date, self.config.max_volume_high_days,
                                                                     max_vol)
        if not is_vol_max_high:
            return None

        # 5天内涨幅不超过xx
        is_five_days_not_more_than = self.data_interface.is_pct_up_not_more_than(day.code, end_date, 5,
                                                                                 self.config.five_days_max_up_pct)
        if not is_five_days_not_more_than:
            return None
        # 10天内涨幅不超过xx
        is_ten_days_not_more_than = self.data_interface.is_pct_up_not_more_than(day.code, end_date, 10,
                                                                                self.config.ten_days_max_up_pct)
        if not is_ten_days_not_more_than:
            return None
        for neg_index in range(index_pos, len(self.daily_lines)):
            if self.daily_lines[neg_index].is_negative():
                if self.daily_lines[neg_index].vol > max_vol:
                    print("neg vol bigger than pos vol")
                    return None

        concept = TushareInterface().get_concept(day.code)

        # concept = None
        name = self.data_interface.get_name(day.code)
        angle_of_5 = TushareInterface().get_slope_of_days(day.code, self.end_date, 5)
        angle_of_10 = TushareInterface().get_slope_of_days(day.code, self.end_date, 10)
        angle_of_20 = TushareInterface().get_slope_of_days(day.code, self.end_date, 20)
        angle_of_30 = TushareInterface().get_slope_of_days(day.code, self.end_date, 30)
        print(angle_of_30)

        is_margined = TushareInterface().is_margin_stock(day.code, previous_daily_line.trade_date[:10].replace("-", ""))
        if self.config.is_margin_stock and not is_margined:
            return None
        searchResult = SearchResult(day.code, name, count, start_date,
                                    end_date, limit_circ_mv, free_circ_mv, concept, max_turnover_rate,
                                    angle_of_5, angle_of_10, angle_of_20, angle_of_30)
        return searchResult

    def save_to_xlsx(self, found_stocks, end_date):
        # 创建一个新的工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "Search Results"

        # 定义表头
        headers = ['股票代码', '股票名称', '连阳天数', '放量上涨日期', '上涨结束日期', '限制流通市值',
                   '自由流通市值', '概念']

        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # 根据choose_date对found_stocks进行降序排序
        found_stocks.sort(key=lambda x: x.end_date, reverse=True)

        # 写入数据
        row_idx = 2
        for result in found_stocks:
            col_idx = 1
            for attr in ['code', 'name', 'count', 'start_date', 'end_date', 'limit_circ_mv',
                         'free_circ_mv', 'concept']:
                value = getattr(result, attr)
                ws.cell(row=row_idx, column=col_idx, value=value)
                col_idx += 1
            row_idx += 1

            # 设置列宽（以字符为单位，可以根据需要调整）
        column_widths = {
            'A': 10,  # 股票代码
            'B': 10,  # 股票名称
            'C': 10,  # 放量上涨日期
            'D': 10,  # 结束日期
            'E': 10,  # limit_circ_mv
            'F': 10,  # free_circ_mv
            'G': 20,  # 所属概念
        }

        for column_letter, width in column_widths.items():
            column = get_column_letter(ws[f'{column_letter}1'].column)
            ws.column_dimensions[column].width = width

        # 保存工作簿
        now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        file_name = f"search_results_{now}.xlsx"
        wb.save(file_name)
        return file_name

    def find(self):
        for index, code in enumerate(self.stock_list):

            self.count += 1
            # print('call stk_factor time:', datetime.now(), self.count)
            # print(code)
            daily_lines = self.data_interface.get_daily_lines(code, self.end_date, self.back_days)
            if daily_lines is None:
                continue
            for index_1 in range(0, len(daily_lines)):
                self.set_daily_lines(daily_lines[index_1:len(daily_lines)])

                result = self.search()
                if result is not None:
                    file.write(f"{code},{result.start_date},{result.end_date}\n")
                    self.search_results.append(result)
                    break

            progress = (index + 1) / len(self.stock_list) * 100
            if index % 20 == 0:
                print(f"Processing progress: {progress:.2f}%")
        return self.search_results

    def update_realtime_data(self, end_date):
        today = self.data_interface.get_today_date()
        res = self.data_interface.is_a_stock_trading_day(
            today)
        res = today == end_date
        res = self.data_interface.is_between_9_30_and_19_00()
        if self.data_interface.is_a_stock_trading_day(
                today) and today == end_date and self.data_interface.is_between_9_30_and_19_00():
            print('start')
            self.data_interface.get_all_realtime_data(self.stock_list)

    def today_codes(self, day):
        today_results = []
        search_results = self.find()
        for result in search_results:
            if result.choose_date == day:
                today_results.append(result)

        # return today_results
        sorted_results = sorted(today_results, key=lambda x: x.score, reverse=True)

        # 检查排序后的列表长度，然后切片以获取前三个元素或全部元素
        top_three = sorted_results[:3]
        return top_three

    def get_month_revenue_csv(self):
        # 获取当前日期
        now = datetime.now()

        # 格式化日期为'YYYYMM'
        current_year_month = now.strftime('%Y%m')
        return f'./revenue/{current_year_month}.csv'
